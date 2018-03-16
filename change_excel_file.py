from openpyxl import load_workbook

test = load_workbook('test.xlsx')
fu_test = load_workbook('2018.1.1.xlsx')
ws_fu_test = fu_test.active

sheet_number = []
for x in range(0,30):
	sheet_number.append(x)
iter_sheet_number = iter(sheet_number)

N2 = []
for i in range(9,42):
	N2.append(i)
iter2 = iter(N2)

rows_fu = ws_fu_test.rows
columns_fu = ws_fu_test.columns
value_fu = []
for row in rows_fu:
	line = [col.value for col in row]
	value_fu.append(line)

	

def write_line():
	sheets = test.get_sheet_names()   
	ws_test = test.get_sheet_by_name(sheets[next(iter_sheet_number)])
	numbers = []
	except_numbers = []
	for i in range(1,150,11):
		except_numbers.append(i)
	for i in range(2,150):
		if i in except_numbers:
			pass
		else:	
			numbers.append(i)
	iter1 = iter(numbers)
	for i in range(2,12):
		for j in range(1,11):
				if value_fu[i][0]==None and value_fu[i][1] == None:
					for q in range(0,j):
						del value_fu[0]
					break
				else:
					ws_test.cell(row = 10, column = next(iter1,151)).value = value_fu[i][j]
write_line()
write_line()
write_line()
write_line()
write_line()
write_line()
write_line()
write_line()

test.save(filename = 'new_test.xlsx')
