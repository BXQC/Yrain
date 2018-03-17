from openpyxl import load_workbook
from openpyxl import Workbook
test = load_workbook('test.xlsx')
fu_test = load_workbook('2018.1.31.xlsx')
ws_fu_test = fu_test.active
#导入编辑模块,并导入本地文件
sheet_number = []
for x in range(0,30):
	sheet_number.append(x)


iter_sheet_number = iter(sheet_number)
#这里是用来迭代每一个要被复制的工作蒲的,暂时没用到.

N2 = []
for i in range(9,42):
	N2.append(i)


iter2 = iter(N2)
#这里用来迭代对应每一个要被复制的工作蒲的位置的行数.暂时没用到.

rows_fu = ws_fu_test.rows
columns_fu = ws_fu_test.columns
value_fu = []
for row in rows_fu:
	line = [col.value for col in row]
	value_fu.append(line)


#这里是用来对导入的要被复制的文件进行读取,读取为list格式.
#此处展示示例:
'''[['监测点  南京', 'AQI', '空气质量', '首要污染物', 'PM2.5', 'PM10', 'CO', 'NO2', 'O3', 'O3', 'SO2'], 
 [None, None, '指数类别', None, '细颗粒物', '可吸入颗粒物', '一氧化碳', '二氧化氮', '臭氧1小时平均', '臭氧8小时平均', '二氧化硫'], 
 ['迈皋桥', 258, '重度污染', '细颗粒物(PM2.5)', 208, 329, 1.4, 82, 69, 56, 11],
 ['草场门', 268, '重度污染', '细颗粒物(PM2.5)', 218, 323, 1.3, 57, 92, 44, 24],
 ['山西路', 239, '重度污染', '细颗粒物(PM2.5)', 189, 336, 0.7, 71, 103, 44, 18], 
 ['中华门', 265, '重度污染', '细颗粒物(PM2.5)', 215, 299, 1.8, 59, 104, 39, 18],
 ['瑞金路', 280, '重度污染', '细颗粒物(PM2.5)', 230, 305, 1.6, 53, 95, 31, 17],
 ['玄武湖', 281, '重度污染', '细颗粒物(PM2.5)', 231, 293, 1.8, 61, 99, 46, 26],
 ['浦口', '_', None, '_', '_', '_', '_', '_', '_', '_', '_'],
 ['奥体中心', 275, '重度污染', '细颗粒物(PM2.5)', 225, 268, 1.3, 64, 25, 16, 17],
 ['仙林大学城', 268, '重度污染', '细颗粒物(PM2.5)', 218, 330, 1.9, 64, 94, 40, 13],
 [None, None, None, None, None, None, None, None, None, None, None],
 ['监测点  无锡', 'AQI', '空气质量', '首要污染物', 'PM2.5', 'PM10', 'CO', 'NO2', 'O3', 'O3', 'SO2'],
 [None, None, '指数类别', None, '细颗粒物', '可吸入颗粒物', '一氧化碳', '二氧化氮', '臭氧1小时平均', '臭氧8小时平均', '二氧化硫'],
 ['曹张', 248, '重度污染', '细颗粒物(PM2.5)', 198, 309, 2.1, 58, 112, 38, 23],'''

#由示例可以得知若将该列表进行迭代,并获取迭代的数据将迭代数据输入对应汇总表相应位置内.

iter_fu_1 = iter(value_fu)
#这里设置了要被复制的工作表的list的迭代对象.

sheets_name = test.get_sheet_names()
s_name = iter(sheets_name)

#这里获取了汇总表的工作表名称,并设置了相应的迭代对象.
#接下来要写将以上数据进行分割,分割成对应工作表的名称的.
#推掉重写,分两个部分实现,第一个部分:对所取的列表的处理,将一个表格的内容完全打包为一行,作为一个单位,也就是共有29个表,有29个list.
#大概用pop来写的话是反过来....除非将原来生成的list就进行一次翻转....
#将原来的list翻转:
list1 = list(value_fu)
list_value_0 = []
for i in range(0,len(value_fu)):
	list_value_0.append(list1.pop)
#这段居然有问题...........依然有问题....先放着...

#list_value_0就是原来输出的列表的翻转,而value_fu == []...空的了..
#现在来写判定什么时候结束什时候开始....这个感觉不好写...先看看没个要被复制的文件之间有没有什么区别...实在不想用里面的值来判定,经过上几次的试验,发现利用迭代器是好,但会出现一些意料之外的BUG...先按着好写的来,假如大部分文件之间数据的位置都一样,就按位置来,不行就真的只能再试一次数据判定来抓取了...
#很好,我看了一下,果然每一个都不一样,...只能用判定来写了,,不过先别迭代了,,就最基础的判定来写....一定可以写的...

'''共有29个组
我们来看一下每个组之间的监测词:
1	迈皋桥-----仙林大学城
2	曹张 -----荣巷
3	黄河新村----铜山区招生办
4	上方山-----相城区
5	市监测站-----钟楼
6	南郊 ------紫琅学院
7	市环境监测站-----德源药业(直管站)
8	钵池山-----淮阴区监测站
9	盐城电厂-----宝龙广场
10	城东财政所----五台山医院
11	环境监测站----市疾控中心
12	公园路----王营
13	市监测站-----市供电局
14	职工医院-----铁路设计院
15	监测站----雷台#这里只需要找到雷台的位置然后向上一格就是第一个监测站
16	监测站----科委#同上
17	仓门街子站----富康家世界子站
18	市科委----公司二招
19	气象局---酒钢宾馆
20	八廓街-----拉萨火车站
21	昌都监测站-----昌都地区昌都坝
22	山南监测站----山南人民医院
23	日喀则监测站----日喀则鸿达公司
24	那曲监测站----那曲会议中心
25	阿里监测站----阿里地委
26	林芝监测站-----林芝人民医院
27	市环境监测站----城北区政府
28	平安县----平安高铁新区
29	海晏县西海镇
'''
#正好是29个组,现在需要写出一个能检测出他们所在行的位置....不过有个很讨打的位置是居然有两个监测站...........................我现在是要将每个单元格进行遍历并且将以上范围的单元个组在一起,最好是成单独的list,
#不能这样写...这样根本监测不到我想要的位置.需要看看有没有其他方法....别急.
name_jiansuo = []
for i in ws_fu_test['A']:
	name_jiansuo.append(i.value)

#这里将要复制的表的A列的所有值作为列表输出了,我们可以用这个检索表来找到我们想要的数字的位置.
#现在写用于导出数值的程序...
a1 = name_jiansuo.index('迈皋桥')#+8
a2 = name_jiansuo.index('曹张')#+7
a3 = name_jiansuo.index('黄河新村')#+8
a4 = name_jiansuo.index('上方山')#+7
a5 = name_jiansuo.index('钟楼')#-8
a6 = name_jiansuo.index('南郊')#+4
a7 = name_jiansuo.index('德源药业(直管站)')#-6
a8 = name_jiansuo.index('钵池山')#+4
a9 = name_jiansuo.index('盐城电厂')#+4
a10 = name_jiansuo.index('城东财政所')#+4
a11 = name_jiansuo.index('环境监测站')#+4
a12 = name_jiansuo.index('公园路')#+4
a13 = name_jiansuo.index('市供电局')#-4
a14 = name_jiansuo.index('职工医院')#+4
a15 = name_jiansuo.index('雷台')#-1
a16 = name_jiansuo.index('科委')#-1
a17 = name_jiansuo.index('仓门街子站')#+1
a18 = name_jiansuo.index('市科委')#+2
a19 = name_jiansuo.index('气象局')#+2
a20 = name_jiansuo.index('八廓街')#+5
a21 = name_jiansuo.index('昌都监测站')#+2
a22 = name_jiansuo.index('山南监测站')#+1
a23 = name_jiansuo.index('日喀则监测站')#+1
a24 = name_jiansuo.index('那曲监测站')#+2
a25 = name_jiansuo.index('阿里监测站')#+1
a26 = name_jiansuo.index('林芝监测站')#+1
a27 = name_jiansuo.index('城北区政府')#-4
a28 = name_jiansuo.index('平安县')#+1
a29 = name_jiansuo.index('海晏县西海镇')



#上面获取了这些位置的所有的数据,但这只是其中一端需要写出另外一端的数据.
#现在写出了所在的位置了,接下来是导出所在位置的值.
s1 = []
for i in range(a1,a1+9):
	for j in range(0,11):
		s1.append(value_fu[i][j])
s2 = []
for i in range(a2,a2+8):
	for j in range(0,11):
		s2.append(value_fu[i][j])
s3 = []
for i in range(a3,a3+9):
	for j in range(0,11):
		s3.append(value_fu[i][j])
s4 = []
for i in range(a4,a4+8):
	for j in range(0,11):
		s4.append(value_fu[i][j])
s5 = []
for i in range(a5-8,a5+1):
	for j in range(0,11):
		s5.append(value_fu[i][j])
s6 = []
for i in range(a6,a6+5):
	for j in range(0,11):
		s6.append(value_fu[i][j])
s7 = []
for i in range(a7-6,a7+1):
	for j in range(0,11):
		s7.append(value_fu[i][j])
s8 = []
for i in range(a8,a8+5):
	for j in range(0,11):
		s8.append(value_fu[i][j])
s9 = []
for i in range(a9,a9+5):
	for j in range(0,11):
		s9.append(value_fu[i][j])
s10= []
for i in range(a10,a10+5):
	for j in range(0,11):
		s10.append(value_fu[i][j])
s11= []
for i in range(a11,a11+5):
	for j in range(0,11):
		s11.append(value_fu[i][j])
s12= []
for i in range(a12,a12+5):
	for j in range(0,11):
		s12.append(value_fu[i][j])
s13= []
for i in range(a13-4,a13+1):
	for j in range(0,11):
		s13.append(value_fu[i][j])
s14= []
for i in range(a14,a14+5):
	for j in range(0,11):
		s14.append(value_fu[i][j])
s15= []
for i in range(a15-1,a15+1):
	for j in range(0,11):
		s15.append(value_fu[i][j])
s16= []
for i in range(a16-1,a16+1):
	for j in range(0,11):
		s16.append(value_fu[i][j])
s17= []
for i in range(a17,a17+2):
	for j in range(0,11):
		s17.append(value_fu[i][j])
s18= []
for i in range(a18,a18+3):
	for j in range(0,11):
		s18.append(value_fu[i][j])
s19= []
for i in range(a19,a19+3):
	for j in range(0,11):
		s19.append(value_fu[i][j])
s20= []
for i in range(a20,a20+6):
	for j in range(0,11):
		s20.append(value_fu[i][j])
s21= []
for i in range(a21,a21+3):
	for j in range(0,11):
		s21.append(value_fu[i][j])
s22= []
for i in range(a22,a22+2):
	for j in range(0,11):
		s22.append(value_fu[i][j])

s23= []
for i in range(a23,a23+2):
	for j in range(0,11):
		s23.append(value_fu[i][j])
s24= []
for i in range(a24,a24+3):
	for j in range(0,11):
		s24.append(value_fu[i][j])
s25= []
for i in range(a25,a25+2):
	for j in range(0,11):
		s25.append(value_fu[i][j])
s26= []
for i in range(a26,a26+1):
	for j in range(0,11):
		s26.append(value_fu[i][j])

s27= []
for i in range(a27-4,a27+1):
	for j in range(0,11):
		s27.append(value_fu[i][j])
s28= []
for i in range(a28,a28+1):
	for j in range(0,11):
		s28.append(value_fu[i][j])
s29= []
for i in range(a29,a29+1):
	for j in range(0,11):
		s29.append(value_fu[i][j])

#现在已近将数据进行了切分,需要的是将数据进行写入,写入的时候应该先读取所有的工作表名称作为要写入的工作表的对应顺序.

#上面有工作表的迭代对象,但我们不使用迭代,因为可能有多个文件要在里面来写,所以我们用循环来写.
#sheet_value = ['s1','s2','s3','s4','s5','s6','s7','s8','s9','s10','s11','s12','s13','s14','s15','s16','s17','s18','s19','s20','s21','s22','s23','s24','s25','s26','s27','s28','s29']
#汇总:
sheet_value = [s1,s2,s3,s4,s5,s6,s7,s8,s9,s10,s11,s12,s13,s14,s15,s16,s17,s18,s19,s20,s21,s22,s23,s24,s25,s26,s27,s28,s29]

sheets = test.get_sheet_names()#这里是获取要被写入的工作表的所有名称的列表

#赋值:
for i in range(0,29):
	sheet_write = test.get_sheet_by_name(sheets[i])
	sheet_write.append(sheet_value[i])
test.save(filename = 'test.xlsx')



















'''def write_sheet_value():
	name = next(s_name)
	check = '监测'
	while True:
		panduan = next(iter_fu_1)
		if (name in panduan[0]):
			pass
		elif (None in panduan[0]):
			pass
		elif (check in panduan[0]) and not(name in panduan[0]):
			break
		else :
			for i in range(1,11):
				ws_fu_test.cell(row = 9,column = i).value = panduan[i]
		return(name)

write_sheet_value()

test.save(filename = 'change_test.xlsx')'''

