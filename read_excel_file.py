from openpyxl import load_workbook

test = load_workbook('test.xlsx')
fu_test = load_workbook('2018.1.1.xlsx')
def write_line():
	sheets = test.get_sheet_names()   
	ws_test = test.get_sheet_by_name(sheets[0])
	ws_fu_test = fu_test.active

	numbers = []
	except_numbers = []


	for i in range(1,150,11):
		except_numbers.append(i)

	for i in range(2,150):
		if i in except_numbers:
			pass
		else:	
			numbers.append(i)

	rows_fu = ws_fu_test.rows
	columns_fu = ws_fu_test.columns
	value_fu = []
	for row in rows_fu:
		line = [col.value for col in row]
		value_fu.append(line)

	iter1 = iter(numbers)


	for i in range(2,12):
		for j in range(1,11):
				if value_fu[i][0]==None:
					for q in range(0,i+1):
						del value_fu[0]
					print(value_fu)
					break
				else:
					ws_test.cell(row = 9, column = next(iter1,151)).value = value_fu[i][j]
	test.save(filename = 'new_test.xlsx')
write_line()