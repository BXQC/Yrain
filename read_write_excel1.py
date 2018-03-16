# -*- coding:  utf-8 -*-
import openpyxl
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
#接下来做一些工作的简便:
wb = openpyxl.workbook(path)#这个是用来管理工作蒲(workbook)的
ws = wb.active#用来访问对应工作蒲的工作表.
ws1 = wb.create_sheet('file_name')#这个用来创建一个新的工作表(wooksheet).
ws.title = 'New Title'#这个是用来改工作表的名称的.
print(wb.sheetnames)#以列表输出文件的工作表的名称.
for sheet in wb:
    print(sheet.title)#这里对工作表做了循坏.
source = wb.active
target = wb.copy_worksheet(source)#这里是将一个工作表整体给复制了一个副本.

c = ws['A4']#c指的是单元格(cell).这里是访问工作表ws的A4单元格的内容.
d = ws.cell(row = 4,column = 2,value = 10)#这个函数可以提供对特定单元格的访问,
            # 其中row(行)column(列)value(值)
for i in range(0,101):
    for j in range(1,101):
        ws.cell(row = i,column = j)#这样做的话会直接在内存里创造出100*100的单元格,并什么都没有做.


cell_range = ws['A1':'C2']#这里是对单元格进行批量访问,使用的是切片的方法.
colC = ws['C']
col_range = ws['C':'D']#这里类似的对行的范围进行了获取
row10 = ws['10']
row_range = ws['10':'14']#这里是对列;
for row in ws.iter_rows(min_row = 2,max_col = 5,max_row = 5):#这里对行和列都选取了范围,并对这个范围内的单元格进行循坏.
    for cell in row:
        print(cell)
for col in ws.iter_cols(min_col = 2,max_row = 5,max_col = 5):
    for cell in col:
        print(cell)
        #这里和上面那个,我觉得没什么大的区别吧,一个是列从0取,另一个是行从0取.这样看来只能选取固定的片了.不能使用这个进行正方形选取了.


    ws = wb.active
    ws['C9'] = 'hello world'
    tuple(ws.rows)#ws.rows用于遍历所有的行和列的单元格.
    tuple(ws.columns)#ws.columns也是用于遍历所有的行和列.
    #区别在于第一个是以行为序列进行输出,第二个是列为序列


    c.value = 'hello world'#通过cell.value进行单元格内容的访问和读写.



    wb.save('balances.xlsx')#这里是将工作蒲储存



    wb = load_workbook('doc.xlsx')
    wb.template = True
    wb.save('doc_template.xlsx')
from openpyxl import load_workbook
    wb = load_workbook('doc_template.xlsx')
    wb.template = False
    wb.save('doc.xlsx',as_template=False)#不太懂这两处,大概一个储存为模块,一个储存为文档...不知道具体做什么.

    wb2 = load_workbook('test.xlsx')
    print(wb2.get_sheet_names())#这个函数用于打开本地已经有的工作蒲.









